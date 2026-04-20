from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from apps.accounts.models import Admin
from apps.attendance.forms import ScanScheduleForm
from django.http import HttpResponse
from openpyxl import Workbook
from .models import ScanSchedule, ScanHistory
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime


def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        admin_id = request.session.get('admin_id')
        if not admin_id:
            return redirect('accounts:login')
        return view_func(request, *args, **kwargs)
    return wrapper


@admin_required 
def dashboard_view(request):
    schedules = ScanSchedule.objects.select_related('admin').all()
    context = {
        'schedules': schedules,
        'admin_name': request.session.get('admin_name', '')
    }
    return render(request, 'attendance/index.html', context)


@admin_required
def create_schedule_view(request):
    form = ScanScheduleForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            admin = Admin.objects.get(id=request.session['admin_id'])
            schedule = form.save(commit=False)
            schedule.admin = admin
            schedule.save()
            messages.success(request, 'Tạo lịch quét thành công.')
            return redirect('attendance:dashboard')

    return render(request, 'attendance/create_schedule.html', {'form': form})


@admin_required
def edit_schedule_view(request, pk):
    schedule = get_object_or_404(ScanSchedule, pk=pk)
    form = ScanScheduleForm(request.POST or None, instance=schedule)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Cập nhật lịch quét thành công.')
            return redirect('attendance:dashboard')

    return render(request, 'attendance/edit_schedule.html', {
        'form': form,
        'schedule': schedule
    })


@admin_required
def toggle_schedule_view(request, pk):
    schedule = get_object_or_404(ScanSchedule, pk=pk)
    schedule.is_active = 0 if schedule.is_active == 1 else 1
    schedule.save(update_fields=['is_active'])
    messages.success(request, 'Đã cập nhật trạng thái lịch quét.')
    return redirect('attendance:dashboard')


@admin_required
def delete_schedule_view(request, pk):
    schedule = get_object_or_404(ScanSchedule, pk=pk)
    schedule.delete()
    messages.success(request, 'Đã xóa lịch quét.')
    return redirect('attendance:dashboard')

@admin_required
def history_list_view(request, pk):
    schedule = get_object_or_404(ScanSchedule, pk=pk)
    histories = ScanHistory.objects.filter(schedule_id=pk).order_by('id')

    return render(request, 'attendance/history_list.html', {
        'schedule': schedule,
        'histories': histories
    })


@admin_required
def export_history_excel_view(request, pk):
    schedule = get_object_or_404(ScanSchedule, pk=pk)
    histories = ScanHistory.objects.filter(schedule_id=pk).order_by('id')

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Danh sach diem danh'

    # Thông tin buổi quét
    sheet['A1'] = 'Chủ đề'
    sheet['B1'] = schedule.topic_name
    sheet['B1'].alignment = Alignment(horizontal='center')

    sheet['A2'] = 'Ngày điểm danh'
    sheet['B2'] = schedule.scan_date
    sheet['B2'].number_format = 'DD/MM/YYYY'
    sheet['B2'].alignment = Alignment(horizontal='center')

    # Header bảng
    sheet.append([])
    sheet.append(['STT', 'Mã sinh viên', 'Tên sinh viên', 'Ngày sinh'])

    header_row = 4
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ghi dữ liệu
    start_row = 5
    for index, item in enumerate(histories, start=1):
        row = start_row + index - 1

        sheet.cell(row=row, column=1, value=index)
        sheet.cell(row=row, column=2, value=item.masv)
        sheet.cell(row=row, column=3, value=item.tensv)

        # Xử lý ngày sinh
        birth_value = item.namsinhsv

        if isinstance(birth_value, str):
            birth_value = birth_value.strip()
            parsed_date = None

            for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d']:
                try:
                    parsed_date = datetime.strptime(birth_value, fmt)
                    break
                except ValueError:
                    continue

            if parsed_date:
                cell = sheet.cell(row=row, column=4, value=parsed_date)
                cell.number_format = 'DD/MM/YYYY'
            else:
                sheet.cell(row=row, column=4, value=birth_value)
        else:
            cell = sheet.cell(row=row, column=4, value=birth_value)
            cell.number_format = 'DD/MM/YYYY'

    # Căn giữa một số cột
    for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=1, max_col=4):
        row[0].alignment = Alignment(horizontal='center')  
        row[1].alignment = Alignment(horizontal='center')  
        row[2].alignment = Alignment(horizontal='center')   
        row[3].alignment = Alignment(horizontal='center')  

    # Chỉnh độ rộng cột
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 18
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'danh_sach_diem_danh_{schedule.id}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook.save(response)
    return response